#!/usr/bin/env python

# This script was originally written by Karl Vandepoele
# It has been modified by Nicolas Vannieuwkerke to make Nextflow implementation easier

############################
### new in version 0.2.8 ###
############################

# deal with empty columns for arriba and starfusion

# TO DO: make script compatible with .gz .vcf files
# always check first if file exists and if not, display error message in the worksheet instead of crashing the script



###########################################################

# importing modules
import argparse
import re
import os
import pandas as pd
import openpyxl
import cyvcf2
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.cell import get_column_letter
from openpyxl.utils import get_column_letter

###############
## constants ##
###############

VCF_INFO_COLUMNS = [
    'CHRA',
    'CHRB',
    'GENEA',
    'GENEB',
    'POSA',
    'POSB',
    'SCORE',
    'TOOL_HITS',
    'SCORE',
    'FRAME_STATUS',
    'TRANSCRIPT_ID_A',
    'TRANSCRIPT_ID_B',
    'EXON_NUMBER_A',
    'EXON_NUMBER_B',
    'TRANSCRIPT_VERSION_A',
    'TRANSCRIPT_VERSION_B',
    'HGNC_ID_A',
    'HGNC_ID_B',
    'ANNOTATIONS',
    'ORIENTATION',
    'FOUND_DB',
    'FOUND_IN',
]

VCF_COLUMNS = [
    "CHROM",
    "POS",
    "ID",
    "REF",
    "ALT",
    "QUAL",
    "FILTER",
    "FORMAT"
]

SCORE_THRESHOLD: float = 0.18
TOOL_HITS_THRESHOLD: int = 1

########################
## defining functions ##
########################

# function to create dir if it does not exist yet
def create_directory(directory_path):
    if not os.path.exists(directory_path):
        os.makedirs(directory_path)
        print(f"Directory '{directory_path}' created.")
    else:
        print(f"Directory '{directory_path}' already exists.")

# Function to fetch needed VCF fields and convert them to a DataFrame
def vcf_to_df(vcf):
    processed_variants: list[dict[str,any]] = []
    for variant in vcf:
        var_dict: dict[str,any] = {}
        var_dict["CHROM"] = update_chroms(variant.CHROM)
        var_dict["POS"] = variant.POS
        var_dict["ID"] = variant.ID
        var_dict["REF"] = variant.REF
        var_dict["ALT"] = update_alt_chroms(variant.ALT)
        var_dict["QUAL"] = variant.QUAL
        var_dict["FILTER"] = variant.FILTER
        var_dict["FORMAT"] = str(variant.FORMAT)
        for column in VCF_INFO_COLUMNS:
            var_dict[column] = variant.INFO.get(column, None)
        var_dict['CHRA'] = update_chroms(var_dict['CHRA'])
        var_dict['CHRB'] = update_chroms(var_dict['CHRB'])
        processed_variants.append(var_dict)
    df = pd.DataFrame(processed_variants)
    if len(df.index) == 0:
        # Create an empty DataFrame with the appropriate headers if the VCF file is empty
        return pd.DataFrame(columns=VCF_INFO_COLUMNS)
    return df

# Adds chr to chromosomes missing it
def update_chroms(chrom:any) -> str:
    if pd.isna(chrom): return chrom
    str_chrom = str(chrom)
    return str_chrom if str_chrom.startswith("chr") else f"chr{str_chrom}"

# Adds chr to alt allele chromosomes missing it
def update_alt_chroms(alts: list[str]) -> (list[str]|str):
    outputs = []
    for alt in alts:
        if "chr" not in alt:
            chr_match = re.match(r"^.*[\[\]]([\dXY]{1,2}):.*$", alt)
            if chr_match:
                chrom:str = chr_match.group(1)
                alt = alt.replace(f"{chrom}:", f"chr{chrom}:")
        outputs.append(alt)
    return outputs[0] if len(outputs) == 1 else outputs

# Function to remove brackets and single quotes
def remove_brackets(df, cols):
    for col in cols:
        df[col] = df[col].apply(lambda x: x[0] if isinstance(x, list) and len(x) == 1 else x)
    return df


def split_string(row, column_name, prefix):
    # check if the cell is a string
    if isinstance(row[column_name], str):
        # split the string into a dictionary
        data = dict(item.split(': ') for item in row[column_name].split(','))
        # add the prefix to the keys
        data = {prefix + key: value for key, value in data.items()}
    else:
        # the cell is not a string, return an empty dictionary (empty values are seen as float by pd)
        data = {}
    # convert the dictionary to a Series and add it to the row
    return pd.concat([row, pd.Series(data, dtype = 'object')])

def create_cell_button(ws, cell_ref, text, link=None,
                        fill='00E8F4F8', font_color='000000',
                        border_color='00B3E5FC', border_style='medium', bold=True):
    c = ws[cell_ref]
    c.value = text
    c.font = Font(color=font_color, bold=bold)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill('solid', fgColor=fill)

    # Medium border for “button” frame
    side = Side(style=border_style, color=border_color)
    c.border = Border(left=side, right=side, top=side, bottom=side)

    # Optional hyperlink
    if link:
        c.hyperlink = link

    # Tidy sizing for a button feel
    ws.row_dimensions[c.row].height = 20
    col_letter = get_column_letter(c.column)
    ws.column_dimensions[col_letter].width = max(len(text) + 6, 18)

def read_arriba_file(arriba_file:str) -> pd.DataFrame:
    df = pd.read_csv(arriba_file, sep='\t')
    df['Fusion'] = df['#gene1'] + "--" + df['gene2']
    df[['str1gene', 'str1fusion']] = df['strand1(gene/fusion)'].str.split('/', n=1,expand=True)
    df[['str2gene', 'str2fusion']] = df['strand2(gene/fusion)'].str.split('/', n=1,expand=True)
    df['ORIENTATION'] = df['str1fusion'] + '/' + df['str2fusion']
    df['TRANSCRIPT_A'] = df['transcript_id1']
    df['TRANSCRIPT_B'] = df['transcript_id2']
    df['FRAME_STATUS'] = df['reading_frame']
    df[['CHRA', 'POSA']] = df['breakpoint1'].str.split(':', n=1, expand=True)
    df[['CHRB', 'POSB']] = df['breakpoint2'].str.split(':', n=1, expand=True)
    df['CHRA'] = df['CHRA'].apply(update_chroms)
    df['CHRB'] = df['CHRB'].apply(update_chroms)
    df['POSA'] = pd.to_numeric(df['POSA'])
    df['POSB'] = pd.to_numeric(df['POSB'])
    df = df[['Fusion', 'CHRA', 'CHRB', 'POSA', 'POSB', 'ORIENTATION', 'TRANSCRIPT_A', 'TRANSCRIPT_B', 'FRAME_STATUS']]
    return df

#############################################
## read in arguments and store as variable ##
#############################################

parser = argparse.ArgumentParser(description="Create a report from output files of the nf-core/rnafusion pipeline")
parser.add_argument('--input', metavar='DIR', type=str, help="Path to a directory containing the input VCFs", required=True)
parser.add_argument('--stringtie', metavar='DIR', type=str, help="Path to a directory containing StringTie gene abundance output files", required=True)
parser.add_argument('--fusionreport', metavar='DIR', type=str, help="Path to a directory containing fusion report files with detected fusion and the amount of reads analysed by each fusion caller", required=True)
parser.add_argument('--ctat', metavar='DIR', type=str, help="Path to a directory containing CTAT-SPLICING cancer intron files", required=True)
parser.add_argument('--multiqc', metavar='DIR', type=str, help="Path to a directory containing the multiQC general stats file", required=True)
parser.add_argument('--output', metavar='DIR', type=str, help="Path to the output directory", required=True)
parser.add_argument('--bams', metavar='DIR', type=str, help="Path to a directory containing BAM files produced by STAR", required=True)
parser.add_argument('--genes', metavar='FILE', type=str, help="Path to a file containing the targeted genes list for the analysis.", required=True)
parser.add_argument('--fusion_whitelist', metavar='FILE', type=str, help="Path to a file containing a fusion whitelist for specific filtering.", required=True)
parser.add_argument('--mane', metavar='FILE', type=str, help="Path to a file containing the MANE transcripts for filtering.", required=True)
parser.add_argument('--run', metavar='STR', type=str, help="Name of the run analysed", required=True)
parser.add_argument('--pipeline_version', metavar='STR', type=str, help="Version of the reporting pipeline used", required=True)
parser.add_argument('--arriba', metavar='STR', type=str, help="Path to arriba output directory", required=True)
parser.add_argument('--design', metavar='STR', type=str, help="TWIST design used in the analysis")

args = parser.parse_args()
input_path:str = args.input + "/"
cov_path:str = args.stringtie + "/"
reads_path:str = args.fusionreport + "/"
splicing_path:str = args.ctat + "/"
qc_path:str = args.multiqc + "/"
output_dir:str = args.output
bam_path:str = args.bams + "/"
genes:str = pd.read_csv(args.genes, sep='\t')
fusions:str = pd.read_csv(args.fusion_whitelist, sep='\t')
mane:str = pd.read_csv(args.mane)
run_nr:str = args.run
pipeline_version:str = args.pipeline_version
arriba_path:str = args.arriba + "/"
design:str = args.design

###############################################################
### Loop over .vcf files, extract info and export to report ###
###############################################################

# extract zipped vcf files or read in zipped format
'''
import gzip

with gzip.open('file.gz', 'rb') as f:
    file_content = f.read()

OR if you want to extract all files:

import os
import gzip
import shutil

for filename in os.listdir(input_path):
    if filename.endswith(".gz"):
        with gzip.open(filename, 'rb') as f_in:
            with open(filename[:-3], 'wb') as f_out:
                shutil.copyfileobj(f_in, f_out)

'''

# Loop over all files in the folder
for filename in os.listdir(input_path):
    # Check if the file is a VCF file
    if filename.endswith(".vcf"):
        # Full path to the VCF file
        vcf_path = os.path.join(input_path, filename)
        vcf = cyvcf2.VCF(vcf_path)

        # Convert VCF to DataFrame
        df = vcf_to_df(vcf)

        # Get sample names
        samples: list[str] = vcf.samples
        if len(samples) != 1:
            raise ValueError(f"Expected exactly one sample in VCF file {filename}, but found {len(samples)} samples.")
        sample_name: str = samples[0]

        print(f"Creating report for sample: {sample_name}")

        # clean up data extracted from info
        # Specify the columns to apply the function to
        # Apply the function to your DataFrame
        df_clean = remove_brackets(df, VCF_INFO_COLUMNS)

        # change annotation of reference sequence to one column
        # output is seen as float, not string (is sometimes empty so pandas consideres it as float)
        df_clean['TRANSCRIPT_A'] = df_clean['TRANSCRIPT_ID_A'].astype(str)
        df_clean = df_clean.drop('TRANSCRIPT_VERSION_A', axis = 1)
        df_clean = df_clean.drop('TRANSCRIPT_ID_A', axis = 1)

        df_clean['TRANSCRIPT_B'] = df_clean['TRANSCRIPT_ID_B'].astype(str)
        df_clean = df_clean.drop('TRANSCRIPT_VERSION_B', axis = 1)
        df_clean = df_clean.drop('TRANSCRIPT_ID_B', axis = 1)

        # Extract the basename and use it as the output Excel file name and worksheet name
        basename = filename.split('_')[0]

        # read info about reads into a new df
        reads_file = os.path.join(reads_path + basename + ".fusions.csv")
        df_reads = pd.read_csv(reads_file)

        # split into different df, depending on algorithm
        df_reads_fc = df_reads.apply(split_string, args=('fusioncatcher', 'fc_'), axis=1)
        df_reads_ar = df_reads.apply(split_string, args=('arriba', 'ar_'), axis=1)
        df_reads_sf = df_reads.apply(split_string, args=('starfusion', 'sf_'), axis=1)

        def define_position(df: pd.DataFrame, column_to_split: str, split_prefix:str, fetch_function: function) -> pd.DataFrame:
            df = df.apply(split_string, args=(column_to_split, split_prefix + '_'), axis=1)
            if f"{split_prefix}_position" in df.columns:
                df[['CHRA', 'POSA', 'CHRB', 'POSB']] = df[f'{split_prefix}_position'][df[f'{split_prefix}_position'].notna()].apply(fetch_function)
                df.drop(columns=[f'{split_prefix}_position', column_to_split], inplace=True)
                df["CHRA"] = df["CHRA"].apply(update_chroms)
                df["CHRB"] = df["CHRB"].apply(update_chroms)
                return df
            else:
                return pd.DataFrame()

        # combine fusion df with reads df on 'Fusion' column
        df_clean['Fusion'] = df_clean['GENEA'] + "--" + df_clean['GENEB']

        merged_df = df_clean
        df_reads_fc = define_position(df_reads, 'fusioncatcher', 'fc', lambda x: pd.Series(re.split(r':\+?-?#?', x)[:-1]))
        if not df_reads_fc.empty:
            merged_df = pd.merge(merged_df, df_reads_fc, on = ['Fusion', 'CHRA', 'CHRB', 'POSA', 'POSB', 'EXON_NUMBER_A', 'EXON_NUMBER_B'], how = 'left')

        df_reads_ar = define_position(df_reads, 'arriba', 'ar', lambda x: pd.Series(re.split(r'[:#]', x)))
        if not df_reads_ar.empty:
            merged_df = pd.merge(merged_df, df_reads_ar, on = ['Fusion', 'CHRA', 'CHRB', 'POSA', 'POSB', 'EXON_NUMBER_A', 'EXON_NUMBER_B'], how = 'left')

        df_reads_sf = define_position(df_reads, 'starfusion', 'sf', lambda x: pd.Series(re.split(r':-?#?', x)[:-1]))
        if not df_reads_sf.empty:
            merged_df = pd.merge(merged_df, df_reads_sf, on = ['Fusion', 'CHRA', 'CHRB', 'POSA', 'POSB', 'EXON_NUMBER_A', 'EXON_NUMBER_B'], how = 'left')

        # drop original fusioncatcher, arriba and starfusion columns
        merged_df = merged_df.drop(merged_df.filter(regex='^(fusioncatcher|arriba|starfusion).*$').columns, axis = 1)

        # drop duplicate columns
        merged_df = merged_df.loc[:, ~merged_df.columns.duplicated()]

        # set certain columns to numerical
        present_columns_numerical: list[str] = list(merged_df.columns.intersection(["POSA", "POSB", 'fc_longest_anchor',
            'fc_common_mapping_reads', 'fc_spanning_pairs','fc_spanning_unique_reads', 'ar_coverage1',
            'ar_coverage2', 'ar_discordant_mates', 'ar_split_reads1', 'ar_split_reads2', 'sf_ffmp',
            'sf_junction_reads', 'sf_spanning_reads']))
        for column in present_columns_numerical:
            merged_df[column] = pd.to_numeric(merged_df[column])

        # limit sf_ffmp to one decimal
        if 'sf_ffmp' in merged_df.columns:
            merged_df['sf_ffmp'] = round(merged_df['sf_ffmp'], 1)

        # add missing annotation data
        ## arriba check
        arriba_nan_df: pd.DataFrame = merged_df[merged_df['FOUND_IN'].str.contains('arriba')]
        arriba_nan_df = arriba_nan_df[(arriba_nan_df["TRANSCRIPT_A"] == 'nan') | (arriba_nan_df["TRANSCRIPT_B"] == 'nan')]
        if not arriba_nan_df.empty:
            arriba_df = read_arriba_file(os.path.join(arriba_path + basename + ".arriba.fusions.tsv"))
            arriba_indexed = arriba_df.set_index(['Fusion', 'CHRA', 'CHRB', 'POSA', 'POSB'])
            merged_indexed = merged_df.set_index(['Fusion', 'CHRA', 'CHRB', 'POSA', 'POSB'])
            merged_df = arriba_indexed.combine_first(merged_indexed).reset_index()

        # subset the dataframe to only contain relevant columns
        relevant_columns: list[str] = [
            "Fusion",
            "CHRA",
            "CHRB",
            "GENEA",
            "GENEB",
            "POSA",
            "POSB",
            "ORIENTATION",
            "FOUND_IN",
            "TOOL_HITS",
            "SCORE",
            "FRAME_STATUS",
            "EXON_NUMBER_A",
            "EXON_NUMBER_B",
            "TRANSCRIPT_A",
            "TRANSCRIPT_B"
        ]

        def starts_with_list(string:str, prefixes:list[str]) -> bool:
            return any(string.startswith(prefix) for prefix in prefixes)

        tool_specific_columns: list[str] = [column for column in merged_df.columns if starts_with_list(column, ['fc_', 'ar_', 'sf_'])]

        relevant_columns.extend(tool_specific_columns)

        df_filt = merged_df[relevant_columns]

        # sort on score
        df_filt = df_filt.sort_values('SCORE', ascending = False)

        # move some columns
        column_to_move = df_filt.pop('EXON_NUMBER_A')
        df_filt.insert(5, 'EXONA', column_to_move)

        column_to_move = df_filt.pop('EXON_NUMBER_B')
        df_filt.insert(6, 'EXONB', column_to_move)

        column_to_move = df_filt.pop('Fusion')
        df_filt.insert(0, 'Fusion', column_to_move)

        df_filt.dropna(subset=tool_specific_columns, how='all', inplace=True)

        # filter base on score and number of callers, both conditions have to be true
        df_final = df_filt[(df_filt['SCORE'] > SCORE_THRESHOLD) & (df_filt['TOOL_HITS'] > TOOL_HITS_THRESHOLD)]

        # ALTERNATIVE: either one of the two conditions is true: gives too many hits
        # df_final = df_filt[(df_filt['SCORE'] > 0.2) | (df_filt['TOOL_HITS'] > 2)]

        # limit filtered to versioned MANE transcript only
        df_MANE = mane['Transcript stable ID version'].astype(str)
        df_final_MANE = df_final[df_final['TRANSCRIPT_A'].apply(lambda x: any(i in x for i in df_MANE)) & df_final['TRANSCRIPT_B'].apply(lambda x: any(i in x for i in df_MANE))]

        #df_final_MANE = df_final[df_final['TRANSCRIPT_A'].isin(df_MANE) | df_final['TRANSCRIPT_B'].isin(df_MANE)]

        #########################
        ### Coverage analysis ###
        #########################

        # read coverage into a new df
        cov_file = os.path.join(cov_path + basename + ".gene.abundance.txt")
        expression = pd.read_csv(cov_file, sep = '\t', dtype={2: 'str'})

        # Round the columns with numerical values to one digit
        columns_to_round = ['Coverage', 'FPKM', 'TPM']
        expression[columns_to_round] = expression[columns_to_round].round(1)

        # filter expression values to contain only targeted genes
        genes_in_df = genes['genes']
        genes_expr = expression[expression['Gene Name'].isin(genes_in_df)]

        # rename column and sort based on gene name
        genes_expr = genes_expr.rename(columns={'Reference': 'Chromosome'})
        genes_expr = genes_expr.sort_values('Gene Name')

        # make separate df for control genes
        ref_genes = ['CHMP2A', 'GPI', 'RAB7A', 'VCP']
        ref_genes_expr = genes_expr[genes_expr['Gene Name'].isin(ref_genes)]

        # make df for whitelisted fusions  (should be placed more upwards in fusion section)
        fusions_in_df = fusions['gene A']
        df_fusions = df_filt[df_filt['GENEA'].isin(fusions_in_df) | df_filt['GENEB'].isin(fusions_in_df)]

        # calculate coverage for DUX4 (mapping quality 0)
        # finaal path nog eens controleren van bashCommand!
        bashCommand = "samtools coverage -r chr4:190173000-190175500 " + bam_path + basename + "*.bam" + " -q 0 -A -H > " + basename + "_DUX4.txt"
        os.system(bashCommand)

        DUX4_file = basename + "_DUX4.txt"

        DUX4_reads = "0"
        if(os.path.exists(DUX4_file)):
            with open(DUX4_file, 'r') as file:
                data = file.read()
                # alternatively look for 'Number of reads' and print these (see separate script for this code)
                result = re.search(r'\((.*?) filtered\)', data)
                if result:
                    DUX4_reads = result.group(1)

        #####################
        ### Splicing data ###
        #####################

        # read in splicing parameters
        splicing_file = os.path.join(splicing_path + basename + ".cancer.introns")
        splicing = pd.read_csv(splicing_file, sep = '\t', dtype={2: 'str'})

        # move variant_name column up front for easy identification
        column_to_move = splicing.pop('variant_name')
        splicing.insert(0, 'variant_name', column_to_move)

        ctat_variants_found: bool = splicing['variant_name'].count()

        ###############
        ### QC data ###
        ###############

        # read in QC parameters
        qc_file = os.path.join(qc_path + "multiqc_general_stats.txt")
        qc = pd.read_csv(qc_file, sep = '\t', dtype={2: 'str'})

        # filter so only the data from this sample are retained
        qc_filt = qc[qc['Sample'].str.contains(basename)]

        def clean_column_name(col_name: str) -> str:
            col_name = col_name.strip().replace('-', ' ').replace('_', ' ')
            return col_name.title()

        qc_filt.columns = [clean_column_name(col) for col in qc_filt.columns]
        # Set Sample as a temporary index to convert all other columns to a float
        qc_filt = qc_filt.set_index(['Sample'], append=True).astype(float).round(2).reset_index(['Sample'])

        ####################################
        ### Save all df to an Excel file ###
        ####################################

        excel_path: str = f"{output_dir}/{sample_name}.xlsx"

        # Sheet names
        fusions_filt_sheet = "fusions_filt"
        fusions_filt_MANE_sheet = "fusions_filt_MANE"
        fusions_specific_sheet = "fusions_specific"
        fusions_all_sheet = "fusions_all"
        splicing_sheet = "CTAT_splicing"
        coverage_ref_sheet = "coverage_ref"
        coverage_all_sheet = "coverage_all"
        qc_sheet = "QC"
        summary_sheet = "Summary"

        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df_final.to_excel(writer, index=False, sheet_name=fusions_filt_sheet)
            df_final_MANE.to_excel(writer, index = False, sheet_name = fusions_filt_MANE_sheet)
            df_fusions.to_excel(writer, index= False, sheet_name=fusions_specific_sheet)
            merged_df.to_excel(writer, index=False, sheet_name=fusions_all_sheet)
            splicing.to_excel(writer, index=False, sheet_name = splicing_sheet)
            ref_genes_expr.to_excel(writer, index= False, sheet_name=coverage_ref_sheet)
            genes_expr.to_excel(writer, index=False, sheet_name=coverage_all_sheet)
            qc_filt.to_excel(writer, index=False, sheet_name=qc_sheet)


        ###############################
        ### Change layout of report ###
        ###############################

        # open report file
        workbook = openpyxl.load_workbook(excel_path)

        # create a summary sheet
        workbook.create_sheet(title=summary_sheet, index=0)
        ws = workbook[summary_sheet]
        ws['A1'] = f"Summary for {basename}"
        ws['A1'].font = Font(size = 14, bold = True)
        ws['A1'].fill = PatternFill("solid", fgColor='00FFCC99')

        ws['A3'] = "Filtered fusions found:"
        ws['B3'] = df_final.shape[0]
        create_cell_button(ws, 'C3', 'Go to data', link=f"#{fusions_filt_sheet}!A1")

        ws['A4'] = "Filtered fusion present in MANE transcripts:"
        ws['B4'] = df_final_MANE.shape[0]
        create_cell_button(ws, 'C4', 'Go to data', link=f"#{fusions_filt_MANE_sheet}!A1")

        ws['A5'] = "Fusions in whitelist found:"
        ws['B5'] = df_fusions.shape[0]
        create_cell_button(ws, 'C5', 'Go to data', link=f"#{fusions_specific_sheet}!A1")

        ws['A6'] = "Total fusions found:"
        ws['B6'] = merged_df.shape[0]
        create_cell_button(ws, 'C6', 'Go to data', link=f"#{fusions_all_sheet}!A1")

        ws['A7'] = 'Splice sites found:'
        ws['B7'] = splicing.shape[0]
        create_cell_button(ws, 'C7', 'Go to data', link=f"#{splicing_sheet}!A1")
        ws['D7'] = f' {ctat_variants_found} CTAT variant{"s" if ctat_variants_found != 1 else ""} found'
        if ctat_variants_found > 0:
            ws['D7'].font = Font(color='00FF0000', bold=True)

        ws['A9'] = "Coverage analysis of control genes:"
        create_cell_button(ws, 'B9', 'Go to data', link=f"#{coverage_ref_sheet}!A1")

        ws['A10'] = "Coverage analysis of all targeted genes:"
        create_cell_button(ws, 'B10', 'Go to data', link=f"#{coverage_all_sheet}!A1")

        ws['A11'] = "QC metrics:"
        create_cell_button(ws, 'B11', 'Go to data', link=f"#{qc_sheet}!A1")

        ws['A13'] = "Patient:"
        ws['A14'] = "RNA-nr:"
        ws['B14'] = basename
        ws['A15'] = "Date of birth:"
        ws['A16'] = "TWIST design:"
        ws['B16'] = design if design != None else "?"
        ws['A17'] = "Sequencing run:"
        ws['B17'] = run_nr

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = len(design) + 5 if design != None and len(design) > 20 else 20
        ws.column_dimensions["D"].width = 25

        # apply styling to columns in summary
        for idx, cell in enumerate(ws['A']):
            if idx == 0: continue  # skip header
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.font = Font(bold=True)

        for cell in ws['B']:
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for cell in ws['D']:
            cell.alignment = Alignment(horizontal='left', vertical='center')

        # modify fusion worksheets
        for worksheet in [fusions_all_sheet, fusions_filt_sheet, fusions_filt_MANE_sheet, fusions_specific_sheet]:
            ws = workbook[worksheet]

            # insert sample name in new empty row and add hyperlink to visualisation files
            ws.insert_rows(1)
            ws['A1'] = "fusions " + basename

            # change layout
            c = ws['A1']
            c.font = Font(size = 14, bold = True)
            c.fill = PatternFill("solid", fgColor='00FFCC99')

            # change column width
            for column_index in range(1, 38):
                column_letter = get_column_letter(column_index)
                ws.column_dimensions[column_letter].width = 12

            for column_index in range(15, 21):
                column_letter = get_column_letter(column_index)
                ws.column_dimensions[column_letter].width = 25

            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["K"].width = 32
            ws.column_dimensions["N"].width = 20

            # change alignment/cell height for header
            for cell in ws["2:2"]:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            ws.row_dimensions[2].height = 50

            # add ensembl link to transcript columns
            def link_transcript(column:str) -> None:
                column_letter: str = openpyxl.utils.cell.get_column_letter(column)
                for cell in ws[f'{column_letter}3:{column_letter}{ws.max_row}']:
                    cell0 = cell[0]
                    if cell0.value not in ["", "nan"]:
                        cell0.hyperlink = f"https://www.ensembl.org/Homo_sapiens/Transcript/Summary?db=core;t={cell0.value}"

            for cell in ws['2']:
                if cell.value == 'TRANSCRIPT_A':
                    link_transcript(cell.column)
                if cell.value == 'TRANSCRIPT_B':
                    link_transcript(cell.column)


        #loop over coverage worksheets to change layout
        for worksheet in [coverage_ref_sheet, coverage_all_sheet]:
            ws = workbook[worksheet]

            # add new line and add sample name
            ws.insert_rows(1)
            ws['A1']= "coverage " + basename

            # change layout
            c = ws['A1']
            c.font = Font(size = 14, bold = True)
            c.fill = PatternFill("solid", fgColor='00FFCC99')

            for column_index in range(1, 10):
                column_letter = get_column_letter(column_index)
                ws.column_dimensions[column_letter].width = 14

            ws.column_dimensions["A"].width = 20

        # add DUX4 data to coverage file
        ws = workbook[coverage_all_sheet]

        ws['J26'] = DUX4_reads + " filtered reads"

        # change layout of splicing worksheet
        for worksheet in [splicing_sheet]:
            ws = workbook[worksheet]

            # add new line and add sample name
            ws.insert_rows(1)
            ws['A1']= "CTAT splicing " + basename

            # change layout
            c = ws['A1']
            c.font = Font(size = 14, bold = True)
            c.fill = PatternFill("solid", fgColor='00FFCC99')

            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 30
            ws.column_dimensions["D"].width = 30
            ws.column_dimensions["E"].width = 20
            ws.column_dimensions["F"].width = 20
            ws.column_dimensions["G"].width = 40
            ws.column_dimensions["H"].width = 40

        # change layout of QC worksheet
        for worksheet in [qc_sheet]:
            ws = workbook[worksheet]

            # add new line and add sample name
            ws.insert_rows(1)
            ws['A1']= "QC " + basename

            # add script version for logging, to be replaced in the future
            ws['E1'] = "nf-cmgg/report version: " + pipeline_version

            # change layout
            c = ws['A1']
            c.font = Font(size = 14, bold = True)
            c.fill = PatternFill("solid", fgColor='00FFCC99')

            ws.column_dimensions["A"].width = 25

            for column_index in range(2, 28):
                column_letter = get_column_letter(column_index)
                ws.column_dimensions[column_letter].width = 15

            for cell in ws["2:2"]:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            ws.row_dimensions[2].height = 100

        # add back to summary button on all sheets except for summary
        for sheet_name in workbook.sheetnames:
            if sheet_name != summary_sheet:
                ws = workbook[sheet_name]
                create_cell_button(ws, 'C1', 'Back to summary', link=f"#{summary_sheet}!A1", border_style='thin', border_color='000000')

        # save file
        workbook.save(excel_path)

